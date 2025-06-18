"""
Custom document loaders for MIDAS3
"""
import os
import csv
import pandas as pd
import xml.etree.ElementTree as ET
from typing import List, Dict, Any, Optional, Union, Iterator, Tuple
from pathlib import Path
from langchain_core.documents import Document
import re
import numpy as np
import warnings
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.merge import MergeCell

class CSVLoader:
    """
    Loader for CSV files that creates documents from rows or entire CSV content.
    Can handle both file paths and file-like objects.
    Enhanced with table formatting and chart data detection capabilities.
    """
    
    def __init__(
        self, 
        file_path: str,
        csv_args: Optional[Dict[str, Any]] = None,
        source_column: Optional[str] = None,
        metadata_columns: Optional[List[str]] = None,
        verbose: bool = True
    ):
        """
        Initialize with file path and optional CSV parsing args.
        
        Args:
            file_path: Path to the CSV file
            csv_args: Dictionary of arguments to pass to pandas.read_csv()
            source_column: Column to use as source (if None, use file path)
            metadata_columns: List of columns to include in metadata
            verbose: Whether to print debug information
        """
        self.file_path = file_path
        self.csv_args = csv_args or {}
        self.source_column = source_column
        self.metadata_columns = metadata_columns or []
        self.verbose = verbose
        
    def _log(self, message: str) -> None:
        """Print log message if verbose mode is enabled."""
        if self.verbose:
            print(f"[CSVLoader] {message}")
            
    def _detect_languages(self, df: pd.DataFrame) -> dict:
        """Detect languages present in the DataFrame.
        
        Args:
            df: DataFrame to analyze for language content
            
        Returns:
            dict: Dictionary mapping language codes to their percentage in the data
        """
        try:
            import langdetect
            from collections import defaultdict
            
            lang_counts = defaultdict(int)
            total_cells = 0
            
            # Sample up to 100 cells to detect languages
            sample_size = min(100, df.size)
            sampled = df.sample(n=min(100, len(df)), axis=0).sample(n=min(100, len(df.columns)), axis=1)
            
            for _, row in sampled.iterrows():
                for cell in row:
                    if pd.isna(cell) or not str(cell).strip():
                        continue
                        
                    try:
                        # Try to detect language
                        lang = langdetect.detect(str(cell))
                        lang_counts[lang] += 1
                        total_cells += 1
                    except:
                        continue
            
            # Calculate percentages
            if total_cells > 0:
                return {lang: (count/total_cells)*100 for lang, count in lang_counts.items()}
            return {}
            
        except ImportError:
            self._log("Warning: langdetect not installed, language detection disabled")
            return {}
    
    def load(self) -> List[Document]:
        """Load CSV file and return documents."""
        try:
            # Validate file path
            if not os.path.isabs(self.file_path):
                self.file_path = os.path.abspath(self.file_path)
            
            if not os.path.exists(self.file_path):
                raise FileNotFoundError(f"CSV file not found at {self.file_path}")
            
            self._log(f"Loading CSV file: {os.path.basename(self.file_path)}")
            
            # Default CSV args if not provided
            default_csv_args = {
                'sep': ',',  # Default to comma, will be overridden if different
                'engine': 'python',  # Use Python engine for better flexibility
                'encoding': 'utf-8-sig',  # handle BOM if present
                'on_bad_lines': 'warn',
                'header': 0,  # use first row as header
                'dtype': str,  # Read all columns as strings to avoid type inference issues
                'skip_blank_lines': True,
                'quotechar': '"',
                'escapechar': '\\',
                'doublequote': True,
                'skipinitialspace': True
            }
            
            csv_args = self.csv_args.copy()
            csv_args.update(default_csv_args)
            
            # First detect the delimiter and read the file
            with open(self.file_path, 'r', encoding='utf-8-sig', errors='replace') as f:
                # Read sample and detect dialect
                sample = f.read(4096)
                try:
                    dialect = csv.Sniffer().sniff(sample.split('\n')[0])
                    # Remove 'sep' if it exists since we'll use 'delimiter'
                    if 'sep' in csv_args:
                        del csv_args['sep']
                    csv_args['delimiter'] = dialect.delimiter
                except Exception as e:
                    self._log(f"Could not detect CSV dialect, using defaults: {e}")
                    # If detection fails, use comma as default and ensure only one of sep/delimiter is set
                    if 'delimiter' in csv_args:
                        del csv_args['delimiter']
                    csv_args['sep'] = ','
                
                # Count total lines for logging
                f.seek(0)
                total_lines = sum(1 for _ in f)
                f.seek(0)
                
                self._log(f"Reading CSV file with {total_lines} lines using delimiter: {repr(csv_args.get('delimiter', csv_args.get('sep', ',')))}")
                
                # Remove engine from kwargs for read_csv
                read_args = {k: v for k, v in csv_args.items() if k != 'engine'}
                try:
                    df = pd.read_csv(f, **read_args)
                    self._log(f"CSV file loaded with {len(df)} rows and {len(df.columns)} columns")
                except Exception as e:
                    self._log(f"Error reading CSV: {e}")
                    # Fallback to basic CSV reading if the first attempt fails
                    f.seek(0)
                    df = pd.read_csv(f, sep=None, engine='python', encoding='utf-8-sig', on_bad_lines='warn')
                    self._log(f"Fallback CSV loading successful: {len(df)} rows and {len(df.columns)} columns")
            
            # If duplicate column names, add suffix to make them unique
            if df.columns.duplicated().any():
                df.columns = [f"{col}_{i}" if i > 0 else col for i, col in enumerate(df.columns)]
            
            # Try to convert date columns
            self._log("Processing date columns...")
            date_cols_count = 0
            for col in df.columns:
                # Try to convert to datetime if it looks like a date
                if df[col].dtype == 'object':
                    try:
                        # Suppress warnings during date parsing
                        with warnings.catch_warnings():
                            warnings.simplefilter("ignore")
                            pd.to_datetime(df[col], errors='raise')
                            df[col] = pd.to_datetime(df[col], errors='coerce')
                            date_cols_count += 1
                    except (ValueError, TypeError):
                        pass
            
            self._log(f"Found {date_cols_count} date columns")
            
            # Initialize metadata with basic file info
            metadata = {
                "source": self.file_path,
                "filename": os.path.basename(self.file_path),
                "file_type": "csv",
                "row_count": len(df),
                "column_count": len(df.columns),
                "columns": ", ".join(df.columns.tolist())
            }
            
            # Detect languages in the data
            lang_stats = self._detect_languages(df)
            if lang_stats:
                # Convert language stats to a string representation for metadata
                lang_str = ', '.join([f'{k}:{v:.1f}%' for k, v in lang_stats.items()])
                metadata["languages"] = lang_str
                self._log(f"Detected languages in CSV: {lang_str}")
                
                # Add primary language as a separate field if available
                if lang_stats:
                    primary_lang = max(lang_stats.items(), key=lambda x: x[1])[0]
                    metadata["primary_language"] = primary_lang
                    self._log(f"Detected primary language in CSV: {primary_lang}")
            
            self._log("Creating documents from CSV data...")
            documents = self._create_documents(df)
            
            self._log(f"Created {len(documents)} documents from CSV data")
            return documents
            
        except Exception as e:
            error_msg = f"Error loading CSV file: {str(e)}"
            self._log(error_msg)
            import traceback
            self._log(traceback.format_exc())
            
            # Return error document with stack trace
            return [Document(
                page_content=error_msg,
                metadata={
                    "source": self.file_path, 
                    "error": str(e),
                    "traceback": traceback.format_exc()
                }
            )]
    
    def _create_documents(self, df: pd.DataFrame) -> List[Document]:
        """Create documents from a pandas DataFrame with enhanced table formatting and chart detection."""
        documents = []
        
        # Detect numeric columns for potential chart data
        numeric_cols = df.select_dtypes(include=['number']).columns.tolist()
        date_cols = [col for col in df.columns if pd.api.types.is_datetime64_dtype(df[col]) or 
                    (df[col].dtype == 'object' and self._is_likely_date_column(df[col]))]
        
        # Generate statistics for numeric columns (only basic stats, no complex objects)
        stats = {}
        if numeric_cols:
            try:
                # Only get basic stats that can be serialized
                stats_df = df[numeric_cols].describe(include='all')
                for col in numeric_cols:
                    col_stats = {}
                    if 'mean' in stats_df[col]:
                        col_stats['mean'] = float(stats_df[col]['mean'])
                    if 'min' in stats_df[col]:
                        col_stats['min'] = float(stats_df[col]['min'])
                    if 'max' in stats_df[col]:
                        col_stats['max'] = float(stats_df[col]['max'])
                    if 'count' in stats_df[col]:
                        col_stats['count'] = int(stats_df[col]['count'])
                    if col_stats:  # Only add if we have stats
                        stats[col] = col_stats
            except Exception as e:
                self._log(f"Warning: Could not generate statistics for numeric columns: {e}")
        
        # Create a document for the entire CSV with better formatting
        # Format as markdown table for better readability
        md_table = self._dataframe_to_markdown(df)
        
        # Combine formats for flexible retrieval
        full_content = f"""# CSV Data: {os.path.basename(self.file_path)}

## Table Format (Markdown):
{md_table}

## Raw Data:
{df.to_string(index=False)}"""
        
        # Enhanced metadata with chart potential indicators
        # Only include simple types that can be serialized
        metadata = {
            "source": self.file_path,
            "type": "full_csv",
            "filename": os.path.basename(self.file_path),
            "rows": int(len(df)),
            "columns": ", ".join(str(col) for col in df.columns.tolist()),
            "numeric_columns": ", ".join(numeric_cols) if numeric_cols else "",
            "date_columns": ", ".join(date_cols) if date_cols else "",
            "has_chart_potential": bool(len(numeric_cols) > 0 and (len(date_cols) > 0 or len(numeric_cols) > 1)),
            "created_at": datetime.now().isoformat()
        }
        
        # Add statistics if available
        if stats:
            for col, col_stats in stats.items():
                # Ensure all values are simple types
                if 'mean' in col_stats:
                    metadata[f"stats_{col}_mean"] = float(col_stats['mean'])
                if 'min' in col_stats:
                    metadata[f"stats_{col}_min"] = float(col_stats['min'])
                if 'max' in col_stats:
                    metadata[f"stats_{col}_max"] = float(col_stats['max'])
        
        # Ensure all metadata values are JSON serializable
        clean_metadata = {}
        for k, v in metadata.items():
            if v is None or isinstance(v, (str, int, float, bool)):
                clean_metadata[k] = v
            else:
                clean_metadata[k] = str(v)
        
        documents.append(Document(
            page_content=full_content,
            metadata=clean_metadata
        ))
        
        # Create documents for each row with paragraph formatting
        for _, row in df.iterrows():
            try:
                # Format row as a paragraph with key-value pairs
                content_parts = []
                for col in df.columns:
                    # Skip empty values
                    if pd.isna(row[col]) or row[col] in ('', None):
                        continue
                        
                    # Format based on data type
                    if pd.api.types.is_numeric_dtype(type(row[col])):
                        content_parts.append(f"{col}: {row[col]:,}")
                    elif pd.api.types.is_datetime64_dtype(type(row[col])):
                        content_parts.append(f"{col}: {row[col].strftime('%Y-%m-%d %H:%M:%S')}")
                    else:
                        content_parts.append(f"{col}: {row[col]}")
                
                # Join with newlines for better readability
                content = "\n".join(content_parts)
                
                # Create enhanced metadata with only simple types
                metadata = {
                    "source": self.file_path,
                    "filename": os.path.basename(self.file_path),
                    "type": "csv_row",
                    "row_index": int(row.name) if row.name is not None else 0,
                    "created_at": datetime.now().isoformat()
                }
                
                # Add selected columns to metadata if specified
                if self.metadata_columns:
                    for col in self.metadata_columns:
                        if col in df.columns and not pd.isna(row[col]):
                            metadata[col] = str(row[col])
                
                # Add numeric values directly to metadata for better querying
                for col in df.columns:
                    if col in numeric_cols and not pd.isna(row[col]):
                        try:
                            metadata[f"numeric_{col}"] = float(row[col])
                        except (ValueError, TypeError) as e:
                            self._log(f"Warning: Could not convert {col} value to float: {e}")
                
                # Ensure all metadata values are JSON serializable
                clean_metadata = {}
                for k, v in metadata.items():
                    if v is None or isinstance(v, (str, int, float, bool)):
                        clean_metadata[k] = v
                    else:
                        clean_metadata[k] = str(v)
                
                documents.append(Document(page_content=content, metadata=clean_metadata))
                
            except Exception as e:
                self._log(f"Error creating document for row {row.name}: {str(e)}")
                continue
        
        return documents
    
    def _is_likely_date_column(self, series: pd.Series) -> bool:
        """Check if a column likely contains date values."""
        # Sample the first few non-null values
        sample = series.dropna().head(5).astype(str)
        if len(sample) == 0:
            return False
            
        # Common date patterns
        date_patterns = [
            r'\d{4}-\d{1,2}-\d{1,2}',  # YYYY-MM-DD
            r'\d{1,2}/\d{1,2}/\d{2,4}',  # MM/DD/YYYY or DD/MM/YYYY
            r'\d{1,2}-\d{1,2}-\d{2,4}',  # MM-DD-YYYY or DD-MM-YYYY
            r'\d{4}/\d{1,2}/\d{1,2}'   # YYYY/MM/DD
        ]
        
        # Check if most values match date patterns
        matches = 0
        for val in sample:
            if any(re.search(pattern, str(val)) for pattern in date_patterns):
                matches += 1
                
        return matches / len(sample) > 0.5
    
    def _dataframe_to_markdown(self, df: pd.DataFrame, max_rows: int = 20) -> str:
        """Convert DataFrame to markdown table format."""
        if len(df) == 0 or len(df.columns) == 0:
            return "*Empty table*"
            
        # Limit rows for large tables
        if len(df) > max_rows:
            display_df = pd.concat([df.head(max_rows//2), df.tail(max_rows//2)])
            ellipsis_row = pd.DataFrame([{col: '...' for col in df.columns}])
            display_df = pd.concat([display_df.head(max_rows//2), ellipsis_row, display_df.tail(max_rows//2)])
        else:
            display_df = df
            
        # Create markdown table
        header = "| " + " | ".join(str(col) for col in display_df.columns) + " |"
        separator = "| " + " | ".join(["---" for _ in display_df.columns]) + " |"
        rows = ["| " + " | ".join(str(cell) for cell in row) + " |" for _, row in display_df.iterrows()]
        
        return "\n".join([header, separator] + rows)
    
    def _dataframe_to_html(self, df: pd.DataFrame, max_rows: int = 50) -> str:
        """Convert DataFrame to HTML table format."""
        try:
            # Limit rows for large tables
            if len(df) > max_rows:
                display_df = pd.concat([df.head(max_rows//2), df.tail(max_rows//2)])
            else:
                display_df = df
                
            # Use pandas HTML formatting
            html = display_df.to_html(index=False, na_rep="N/A", border=1)
            return html
        except Exception as e:
            return f"<p>Error generating HTML table: {str(e)}</p>"


class EnhancedExcelLoader:
    """
    Advanced loader for Excel files with enhanced structure detection and processing.
    Features:
    - Merged cell handling
    - Multi-level header detection
    - Table boundary detection
    - Enhanced date/time normalization
    - Improved metadata extraction for charts and formulas
    - Progress reporting
    """
    
    def __init__(
        self, 
        file_path: str,
        sheet_names: Optional[List[str]] = None,
        source_column: Optional[str] = None,
        metadata_columns: Optional[Dict[str, List[str]]] = None,
        detect_merged_cells: bool = True,
        detect_multi_level_headers: bool = True,
        detect_tables: bool = True,
        extract_formulas: bool = True,
        max_header_rows: int = 3,
        min_table_rows: int = 3,
        verbose: bool = True
    ):
        """
        Initialize with file path and optional Excel parsing args.
        
        Args:
            file_path: Path to the Excel file
            sheet_names: List of sheet names to process (if None, process all sheets)
            source_column: Column to use as source (if None, use file path)
            metadata_columns: Dictionary mapping sheet names to lists of columns to include in metadata
            detect_merged_cells: Whether to detect and handle merged cells
            detect_multi_level_headers: Whether to detect multi-level headers
            detect_tables: Whether to detect logical tables within sheets
            extract_formulas: Whether to extract formula information
            max_header_rows: Maximum number of rows to consider as headers
            min_table_rows: Minimum number of rows to consider a table
            verbose: Whether to print detailed progress information
        """
        # Path validation will be done in load() method
        self.file_path = file_path
        self.sheet_names = sheet_names
        self.source_column = source_column
        self.metadata_columns = metadata_columns or {}
        
        # Enhanced options
        self.detect_merged_cells = detect_merged_cells
        self.detect_multi_level_headers = detect_multi_level_headers
        self.detect_tables = detect_tables
        self.extract_formulas = extract_formulas
        self.max_header_rows = max_header_rows
        self.min_table_rows = min_table_rows
        self.verbose = verbose
        
        # Suppress specific pandas warnings
        warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')
        warnings.filterwarnings('ignore', category=DeprecationWarning)
        warnings.filterwarnings('ignore', category=FutureWarning)
    
    def _log(self, message: str) -> None:
        """Print log message if verbose mode is enabled."""
        if self.verbose:
            print(f"[EnhancedExcelLoader] {message}")
            
    def _get_merged_cell_ranges(self, worksheet) -> Dict[Tuple[int, int], Tuple[int, int]]:
        """
        Get dictionary of merged cell ranges in the worksheet.
        Returns a dict mapping (row, col) to (row_span, col_span).
        """
        merged_cells = {}
        
        for merged_range in worksheet.merged_cells.ranges:
            min_row, min_col = merged_range.min_row, merged_range.min_col
            max_row, max_col = merged_range.max_row, merged_range.max_col
            row_span = max_row - min_row + 1
            col_span = max_col - min_col + 1
            
            # Store the top-left cell and its span
            merged_cells[(min_row, min_col)] = (row_span, col_span)
            
        return merged_cells
        
    def _detect_multi_level_headers(self, df: pd.DataFrame, worksheet) -> Tuple[pd.DataFrame, int]:
        """
        Detect and process multi-level headers in the worksheet.
        Returns the DataFrame with proper multi-level headers and the number of header rows.
        """
        if not self.detect_multi_level_headers:
            return df, 1
            
        # Check for potential header rows (up to max_header_rows)
        max_check = min(self.max_header_rows, len(df))
        if max_check <= 1:
            return df, 1
            
        # Get merged cell information
        merged_cells = self._get_merged_cell_ranges(worksheet) if self.detect_merged_cells else {}
        
        # Check for empty rows that might separate headers from data
        empty_rows = []
        for i in range(max_check):
            if df.iloc[i].isna().all() or df.iloc[i].astype(str).str.strip().eq('').all():
                empty_rows.append(i)
        
        # Check for merged cells in the header area
        header_merged_cells = []
        for (row, col), (row_span, col_span) in merged_cells.items():
            if row <= max_check and row_span > 1:
                header_merged_cells.append((row, row_span))
        
        # Determine header rows based on empty rows and merged cells
        header_rows = 1  # Default
        
        if empty_rows:
            # If there's an empty row, headers might be above it
            header_rows = min(empty_rows) + 1
        elif header_merged_cells:
            # If there are merged cells in the header, use their span
            max_merged_row = max(row + span - 1 for row, span in header_merged_cells)
            header_rows = max(header_rows, max_merged_row)
        
        # Limit header rows to max_header_rows
        header_rows = min(header_rows, self.max_header_rows)
        
        if header_rows > 1:
            # Create multi-level header
            header_df = df.iloc[:header_rows].copy()
            data_df = df.iloc[header_rows:].copy()
            
            # Fill NaN values in header with values from previous rows (for merged cells)
            for i in range(header_rows):
                for j in range(len(header_df.columns)):
                    if pd.isna(header_df.iloc[i, j]) or header_df.iloc[i, j] == '':
                        # Look for a value in previous columns if this is a merged cell that spans columns
                        for k in range(j-1, -1, -1):
                            cell_key = (i+1, k+1)  # Convert to 1-indexed for openpyxl
                            if cell_key in merged_cells:
                                row_span, col_span = merged_cells[cell_key]
                                if k+col_span > j:  # This merged cell covers our column
                                    header_df.iloc[i, j] = header_df.iloc[i, k]
                                    break
            
            # Create proper multi-level columns
            columns = pd.MultiIndex.from_arrays([header_df.iloc[i] for i in range(header_rows)])
            data_df.columns = columns
            
            return data_df, header_rows
        
        return df, 1
        
    def _detect_tables(self, df: pd.DataFrame) -> List[pd.DataFrame]:
        """
        Detect logical tables within a DataFrame based on empty rows and structure.
        Returns a list of DataFrames, each representing a logical table.
        """
        if not self.detect_tables or len(df) < self.min_table_rows * 2:
            return [df]  # Return the entire DataFrame as a single table
            
        # Find empty rows that might separate tables
        empty_row_indices = []
        for i in range(len(df)):
            if df.iloc[i].isna().all() or df.iloc[i].astype(str).str.strip().eq('').all():
                empty_row_indices.append(i)
                
        if not empty_row_indices:
            return [df]  # No empty rows, so treat as a single table
            
        # Add start and end indices to create complete ranges
        table_boundaries = [-1] + empty_row_indices + [len(df)]
        
        # Extract tables between empty rows
        tables = []
        for i in range(len(table_boundaries) - 1):
            start_idx = table_boundaries[i] + 1
            end_idx = table_boundaries[i + 1]
            
            if end_idx - start_idx >= self.min_table_rows:
                table_df = df.iloc[start_idx:end_idx].copy().reset_index(drop=True)
                tables.append(table_df)
                
        # If no tables were found (all too small), return the original DataFrame
        if not tables:
            return [df]
            
        return tables
        
    def _extract_formulas(self, worksheet) -> Dict[str, str]:
        """
        Extract formulas from the worksheet.
        Returns a dictionary mapping cell references to formulas.
        """
        if not self.extract_formulas:
            return {}
            
        formulas = {}
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.data_type == 'f':  # Formula cell
                    cell_ref = f"{get_column_letter(cell.column)}{cell.row}"
                    formulas[cell_ref] = str(cell.value)
                    
        return formulas
        
    def _enhanced_date_detection(self, df: pd.DataFrame) -> Tuple[pd.DataFrame, List[str]]:
        """
        Enhanced date column detection and normalization.
        Returns the DataFrame with normalized dates and a list of date column names.
        """
        date_columns = []
        
        # Common date patterns (more comprehensive than before)
        date_patterns = [
            r'\d{4}-\d{1,2}-\d{1,2}',                # YYYY-MM-DD
            r'\d{1,2}/\d{1,2}/\d{2,4}',              # MM/DD/YYYY or DD/MM/YYYY
            r'\d{1,2}-\d{1,2}-\d{2,4}',              # MM-DD-YYYY or DD-MM-YYYY
            r'\d{4}/\d{1,2}/\d{1,2}',                # YYYY/MM/DD
            r'\d{1,2}\s+[A-Za-z]{3,}\s+\d{2,4}',     # DD Month YYYY
            r'[A-Za-z]{3,}\s+\d{1,2},?\s+\d{2,4}',   # Month DD, YYYY
            r'\d{1,2}\.\d{1,2}\.\d{2,4}'            # DD.MM.YYYY or MM.DD.YYYY
        ]
        
        # Check each column for date patterns
        for col in df.columns:
            col_name = col
            if isinstance(col, tuple):  # Handle multi-level columns
                col_name = ' '.join([str(c) for c in col if str(c).strip()])
            
            # Skip columns that are already datetime
            if df[col].dtype == 'datetime64[ns]':
                date_columns.append(col)
                continue
                
            # Only check string and object columns
            if df[col].dtype not in ['object', 'string']:
                continue
                
            # Sample non-null values
            sample = df[col].dropna().astype(str).sample(min(50, len(df[col].dropna()))).tolist()
            if not sample:
                continue
                
            # Check if most values match date patterns
            matches = 0
            for val in sample:
                if any(re.search(pattern, str(val)) for pattern in date_patterns):
                    matches += 1
                    
            if matches / len(sample) > 0.5:  # More than 50% match date patterns
                date_columns.append(col)
                
                # Try to convert to datetime
                try:
                    df[col] = pd.to_datetime(df[col], errors='coerce', infer_datetime_format=True)
                except Exception as e:
                    self._log(f"Warning: Failed to convert column '{col_name}' to datetime: {str(e)}")
                    
        return df, date_columns
    
    def _detect_languages(self, df: pd.DataFrame) -> dict:
        """Detect languages present in the DataFrame."""
        import langdetect
        from collections import defaultdict
        
        lang_counts = defaultdict(int)
        total_cells = 0
        
        # Sample up to 100 cells to detect languages
        sample_size = min(100, df.size)
        sampled = df.sample(n=min(100, len(df)), axis=0).sample(n=min(100, len(df.columns)), axis=1)
        
        for _, row in sampled.iterrows():
            for cell in row:
                if pd.isna(cell) or not str(cell).strip():
                    continue
                    
                try:
                    # Try to detect language
                    lang = langdetect.detect(str(cell))
                    lang_counts[lang] += 1
                    total_cells += 1
                except:
                    continue
        
        # Calculate percentages
        if total_cells > 0:
            return {lang: (count/total_cells)*100 for lang, count in lang_counts.items()}
        return {}
    
    def load(self) -> List[Document]:
        """Load data from the Excel file and return a list of Documents."""
        try:
            # Validate file path
            if not os.path.isabs(self.file_path):
                self.file_path = os.path.abspath(self.file_path)
            
            if not os.path.exists(self.file_path):
                raise FileNotFoundError(f"Excel file not found at {self.file_path}")
            
            self._log(f"Using enhanced Excel loader for {os.path.basename(self.file_path)}")
            
            # Suppress all pandas warnings
            with warnings.catch_warnings():
                warnings.simplefilter("ignore")
                
                # Get all sheet names
                self._log(f"Reading Excel file structure {os.path.basename(self.file_path)}...")
                
                # Load workbook with openpyxl for advanced features
                workbook = load_workbook(self.file_path, data_only=False)
                available_sheets = workbook.sheetnames
                
                # Filter sheets if specified
                if self.sheet_names:
                    sheets_to_process = [s for s in self.sheet_names if s in available_sheets]
                    if not sheets_to_process:
                        self._log(f"Warning: None of the specified sheets {self.sheet_names} found in {self.file_path}")
                        sheets_to_process = available_sheets
                else:
                    sheets_to_process = available_sheets
                    
                self._log(f"Found {len(sheets_to_process)} sheets to process: {', '.join(sheets_to_process)}")

                # Initialize document collection
                documents = []
                
                # Create a document for the entire Excel file with basic metadata
                file_metadata = {
                    "source": self.file_path,
                    "file_type": "excel",
                    "filename": os.path.basename(self.file_path),
                    "sheet_count": len(sheets_to_process),
                    "sheet_names": sheets_to_process,
                    "creation_date": datetime.now().isoformat(),
                }
                
                file_content = f"Excel file with {len(sheets_to_process)} sheets: {', '.join(sheets_to_process)}"
                documents.append(Document(page_content=file_content, metadata=file_metadata))
                
                # Process each sheet
                for sheet_name in sheets_to_process:
                    self._log(f"Processing sheet: {sheet_name}")
                    
                    # Get the worksheet for advanced processing
                    worksheet = workbook[sheet_name]
                    
                    # Extract formulas if enabled
                    formulas = self._extract_formulas(worksheet)
                    
                    # Read the sheet into a DataFrame with robust handling of number formats
                    try:
                        # First try reading with openpyxl directly for better format handling
                        wb = load_workbook(self.file_path, data_only=True, read_only=True)
                        ws = wb[sheet_name]
                        
                        # Convert worksheet to list of lists
                        data = []
                        for row in ws.iter_rows(values_only=True):
                            data.append(list(row))
                        
                        # Close the workbook to free resources
                        wb.close()
                        
                        # Convert to DataFrame
                        if data:
                            # Find the max row length to handle irregular data
                            max_len = max(len(row) for row in data)
                            # Pad rows with None to ensure consistent length
                            data = [row + [None] * (max_len - len(row)) for row in data]
                            df = pd.DataFrame(data[1:], columns=data[0] if data else None)
                        else:
                            df = pd.DataFrame()
                        
                        # If the above fails, fall back to pandas with error handling
                        if df.empty:
                            self._log("Warning: Openpyxl read produced empty DataFrame, falling back to pandas")
                            df = pd.read_excel(
                                self.file_path,
                                sheet_name=sheet_name,
                                engine='openpyxl',
                                header=None,
                                na_filter=True,
                                keep_default_na=True,
                                dtype=str  # Read everything as string to avoid format issues
                            )
                    except Exception as e:
                        self._log(f"Error reading sheet {sheet_name}: {str(e)}")
                        continue
                        
                    if df.empty:
                        self._log(f"Sheet {sheet_name} is empty, skipping")
                        continue
                        
                    # Apply multi-level header detection
                    df, header_rows = self._detect_multi_level_headers(df, worksheet)
                    
                    # Debug: Print sample data to check content
                    self._log(f"Sample data from {sheet_name} (first 3 rows):")
                    for i, row in df.head(3).iterrows():
                        row_preview = " | ".join([f"{k}:{v}" for k, v in row.items()][:5])  # First 5 columns
                        self._log(f"  Row {i}: {row_preview}...")
                    
                    # Apply enhanced date detection
                    df, date_columns = self._enhanced_date_detection(df)
                    
                    # Detect logical tables within the sheet
                    tables = self._detect_tables(df)
                    
                    # Create metadata for the sheet
                    sheet_metadata = {
                        "source": self.file_path,
                        "sheet_name": sheet_name,
                        "row_count": len(df),
                        "column_count": len(df.columns),
                        "multi_level_header": header_rows > 1,
                        "header_rows": header_rows,
                        "date_columns": [str(col) for col in date_columns],
                        "has_formulas": len(formulas) > 0,
                        "formula_count": len(formulas),
                        "table_count": len(tables),
                        "merged_cells": len(self._get_merged_cell_ranges(worksheet)) > 0,
                        "languages_present": self._detect_languages(df),  # Add language detection
                    }
                    
                    # Debug: Log metadata
                    self._log(f"Sheet metadata: {json.dumps(sheet_metadata, indent=2, default=str)}")
                    
                    # Add metadata columns if specified
                    if sheet_name in self.metadata_columns:
                        for col in self.metadata_columns[sheet_name]:
                            if col in df.columns:
                                col_values = df[col].dropna().astype(str).tolist()
                                if col_values:
                                    sheet_metadata[f"meta_{col}"] = col_values
                    
                    # Generate content for the sheet document
                    sheet_content = f"Sheet: {sheet_name}\n\n"
                    
                    # Add table summary
                    sheet_content += f"Contains {len(tables)} logical table(s)\n"
                    
                    # Add sample data in markdown format
                    for i, table_df in enumerate(tables):
                        if i > 0:
                            sheet_content += "\n---\n\n"  # Separator between tables
                        sheet_content += f"Table {i+1} ({len(table_df)} rows x {len(table_df.columns)} columns):\n\n"
                        sheet_content += self._dataframe_to_markdown(table_df)
                    
                    # Create document for the sheet
                    documents.append(Document(page_content=sheet_content, metadata=sheet_metadata))
                    
                    # Create row documents for each table if there are enough rows
                    for table_idx, table_df in enumerate(tables):
                        if len(table_df) > 5:  # Only create row documents for tables with enough rows
                            # Create a document for each row
                            for row_idx, row in table_df.iterrows():
                                row_content = f"Row {row_idx+1} from sheet '{sheet_name}', table {table_idx+1}:\n\n"
                                
                                # Format row data
                                for col in table_df.columns:
                                    col_name = col
                                    if isinstance(col, tuple):
                                        col_name = ' '.join([str(c) for c in col if str(c).strip()])
                                    row_content += f"{col_name}: {row[col]}\n"
                                
                                # Create row metadata
                                row_metadata = {
                                    "source": self.file_path,
                                    "sheet_name": sheet_name,
                                    "table_index": table_idx,
                                    "row_index": row_idx,
                                    "document_type": "excel_row",
                                }
                                
                                # Add source column to metadata if specified
                                if self.source_column and self.source_column in table_df.columns:
                                    row_metadata["source_value"] = str(row[self.source_column])
                                
                                documents.append(Document(page_content=row_content, metadata=row_metadata))
                
                self._log(f"Processed {len(sheets_to_process)} sheets, created {len(documents)} documents")
                return documents
                
        except FileNotFoundError as e:
            raise e
        except Exception as e:
            self._log(f"Error processing Excel file: {str(e)}")
            # Return a basic error document
            error_metadata = {
                "source": self.file_path,
                "error": str(e),
                "error_type": type(e).__name__,
            }
            error_content = f"Error processing Excel file {os.path.basename(self.file_path)}: {str(e)}"
            return [Document(page_content=error_content, metadata=error_metadata)]
    
    def _dataframe_to_markdown(self, df: pd.DataFrame, max_rows: int = 20) -> str:
        """Convert DataFrame to markdown table format."""
        if len(df) == 0 or len(df.columns) == 0:
            return "*Empty table*"
            
        # Limit rows for large tables
        if len(df) > max_rows:
            display_df = pd.concat([df.head(max_rows//2), df.tail(max_rows//2)])
            ellipsis_row = pd.DataFrame([{col: '...' for col in df.columns}])
            display_df = pd.concat([display_df.head(max_rows//2), ellipsis_row, display_df.tail(max_rows//2)])
        else:
            display_df = df
            
        # Handle multi-level columns
        if isinstance(df.columns, pd.MultiIndex):
            header_rows = []
            for level in range(df.columns.nlevels):
                header_row = [str(col[level]) if isinstance(col, tuple) else str(col) for col in df.columns]
                header_rows.append("| " + " | ".join(header_row) + " |")
            header = "\n".join(header_rows)
        else:
            # Create single-level header
            header = "| " + " | ".join(str(col) for col in display_df.columns) + " |"
            
        separator = "| " + " | ".join(["---" for _ in display_df.columns]) + " |"
        
        # Format rows, handling potential multi-index columns
        rows = []
        for _, row_data in display_df.iterrows():
            row_values = []
            for col in display_df.columns:
                val = row_data[col]
                # Format dates nicely
                if pd.api.types.is_datetime64_any_dtype(val):
                    val = val.strftime('%Y-%m-%d') if not pd.isna(val) else ''
                row_values.append(str(val))
            rows.append("| " + " | ".join(row_values) + " |")
        
        return "\n".join([header, separator] + rows)
                
class ExcelLoader:
    """
    Loader for Excel files that creates documents from sheets or entire Excel file.
    Enhanced with table formatting and chart data detection capabilities.
    """
    
    def __init__(
        self, 
        file_path: str,
        sheet_names: Optional[List[str]] = None,
        source_column: Optional[str] = None,
        metadata_columns: Optional[Dict[str, List[str]]] = None
    ):
        """
        Initialize with file path and optional Excel parsing args.
        
        Args:
            file_path: Path to the Excel file
            sheet_names: List of sheet names to process (if None, process all sheets)
            source_column: Column to use as source (if None, use file path)
            metadata_columns: Dictionary mapping sheet names to lists of columns to include in metadata
        """
        # Path validation will be done in load() method
        
        self.file_path = file_path
        self.sheet_names = sheet_names
        self.source_column = source_column
        self.metadata_columns = metadata_columns or {}
    
    def load(self) -> List[Document]:
        """Load data from the Excel file and return a list of Documents."""
        try:
            # Validate file path
            if not os.path.isabs(self.file_path):
                self.file_path = os.path.abspath(self.file_path)
            
            if not os.path.exists(self.file_path):
                raise FileNotFoundError(f"Excel file not found at {self.file_path}")
            
            print(f"Using custom Excel loader for {os.path.basename(self.file_path)}")
            
            # Suppress all pandas warnings
            with warnings.catch_warnings():
                warnings.simplefilter("ignore")
                
                # Get all sheet names
                print(f"Reading Excel file structure {os.path.basename(self.file_path)}...")
                xls = pd.ExcelFile(self.file_path, engine='openpyxl')
                sheet_names = self.sheet_names if self.sheet_names else xls.sheet_names
                total_rows = 0
                
                for sheet_name in sheet_names:
                    try:
                        # Read the sheet with first row as header
                        df = pd.read_excel(
                            xls,
                            sheet_name=sheet_name,
                            engine='openpyxl',
                            header=0  # Use first row as header
                        )
                        
                        # If duplicate column names, add suffix to make them unique
                        if df.columns.duplicated().any():
                            df.columns = [f"{col}_{i}" if i > 0 else col for i, col in enumerate(df.columns)]
                        
                        print(f"Sheet '{sheet_name}' loaded with {len(df)} rows and {len(df.columns)} columns")
                        total_rows += len(df)
                        
                        # Try to convert date columns
                        date_cols_count = 0
                        print(f"Processing date columns in sheet '{sheet_name}'...")
                        
                        # Check for duplicate columns again after potential modifications
                        if df.columns.duplicated().any():
                            df.columns = [f"{col}_{i}" if i > 0 else col for i, col in enumerate(df.columns)]
                        
                        # Try to detect date columns
                        for col in df.columns:
                            if df[col].dtype == 'object':
                                try:
                                    with warnings.catch_warnings():
                                        warnings.simplefilter("ignore")
                                        pd.to_datetime(df[col], errors='raise')
                                        df[col] = pd.to_datetime(df[col], errors='coerce')
                                        date_cols_count += 1
                                except (ValueError, TypeError):
                                    pass
                        
                        print(f"Found {date_cols_count} date columns in sheet '{sheet_name}'")
                        print(f"Creating documents from sheet '{sheet_name}'...")
                        
                        # Create documents for this sheet
                        sheet_documents = self._create_sheet_documents(df, sheet_name)
                        print(f"Created {len(sheet_documents)} documents from sheet '{sheet_name}'")
                        all_documents.extend(sheet_documents)
                    except Exception as e:
                        print(f"Error processing sheet '{sheet_name}': {str(e)}")
                        continue
                
                print(f"Excel processing complete: {len(all_documents)} total documents created from {len(sheet_names)} sheets with {total_rows} total rows")
                return all_documents
            
            return []
            
        except Exception as e:
            return [Document(
                page_content=f"Error loading Excel file: {str(e)}",
                metadata={"source": self.file_path, "error": str(e)}
            )]
    

        
        # Also check for string columns that might be numeric but parsed as strings
        for col in df.select_dtypes(include=['object']).columns:
            try:
                # Check if column can be converted to numeric
                pd.to_numeric(df[col], errors='raise')
                if col not in numeric_cols:
                    numeric_cols.append(col)
            except (ValueError, TypeError):
                pass
                
        # Detect date columns
        date_cols = [col for col in df.columns if pd.api.types.is_datetime64_dtype(df[col]) or 
                    (df[col].dtype == 'object' and self._is_likely_date_column(df[col]))]
        
        # Generate statistics for numeric columns
        stats = {}
        if numeric_cols:
            try:
                stats = df[numeric_cols].describe().to_dict()
            except:
                pass
        
        # Create a document for the entire sheet with better formatting
        md_table = self._dataframe_to_markdown(df)
        
        # Create paragraph-formatted content for each record
        records_content = []
        for _, row in df.iterrows():
            record_parts = []
            for col in df.columns:
                # Skip empty values
                if pd.isna(row[col]) or row[col] in ('', None):
                    continue
                    
                # Format based on data type
                if col in numeric_cols:
                    record_parts.append(f"{col}: {row[col]:,}")
                elif col in date_cols:
                    if pd.api.types.is_datetime64_dtype(type(row[col])):
                        record_parts.append(f"{col}: {row[col].strftime('%Y-%m-%d %H:%M:%S')}")
                    else:
                        record_parts.append(f"{col}: {row[col]}")
                else:
                    record_parts.append(f"{col}: {row[col]}")
            
            if record_parts:  # Only add non-empty records
                records_content.append("\n".join(record_parts) + "\n" + ("-" * 50))
        
        # Combine all content
        full_content = f"""# Excel Sheet: {sheet_name} from {os.path.basename(self.file_path)}

## Table Format (Markdown):
{md_table}

## Records:
""" + "\n\n".join(records_content)
        
        # Enhanced metadata with chart potential indicators
        metadata = {
            "source": self.file_path,
            "type": "excel_sheet",
            "sheet": sheet_name,
            "rows": len(df),
            "columns": ", ".join(df.columns.tolist()),
            "numeric_columns": ", ".join(numeric_cols) if numeric_cols else "",
            "date_columns": ", ".join(date_cols) if date_cols else "",
            "has_chart_potential": len(numeric_cols) > 0 and (len(date_cols) > 0 or len(numeric_cols) > 1),
            "html_table": html_table
        }
        
        # Add statistics if available
        if stats:
            for col, col_stats in stats.items():
                # Add min, max, mean for each numeric column
                if 'mean' in col_stats:
                    metadata[f"stats_{col}_mean"] = col_stats['mean']
                if 'min' in col_stats:
                    metadata[f"stats_{col}_min"] = col_stats['min']
                if 'max' in col_stats:
                    metadata[f"stats_{col}_max"] = col_stats['max']
        
        documents.append(Document(
            page_content=full_content,
            metadata=metadata
        ))
        
        # Get metadata columns for this sheet
        sheet_metadata_cols = self.metadata_columns.get(sheet_name, [])
        
        # Create documents for each row with enhanced formatting
        for _, row in df.iterrows():
            # Format row as key-value pairs with better formatting
            content_parts = []
            for col in df.columns:
                # Format based on data type
                if pd.api.types.is_numeric_dtype(type(row[col])):
                    content_parts.append(f"{col}: {row[col]:,}")
                else:
                    content_parts.append(f"{col}: {row[col]}")
            
            content = "\n".join(content_parts)
            
            # Create enhanced metadata
            metadata = {
                "source": self.file_path,
                "type": "excel_row",
                "sheet": sheet_name,
                "row_index": int(row.name) if row.name is not None else 0
            }
            
            # Add selected columns to metadata if specified
            if sheet_metadata_cols:
                for col in sheet_metadata_cols:
                    if col in df.columns:
                        metadata[col] = str(row[col])
            
            # Add numeric values directly to metadata for better querying
            for col in df.columns:
                if col in numeric_cols:
                    try:
                        metadata[f"numeric_{col}"] = float(row[col])
                    except (ValueError, TypeError):
                        pass
            
            documents.append(Document(page_content=content, metadata=metadata))
        
        return documents
        
    def _is_likely_date_column(self, series: pd.Series) -> bool:
        """Check if a column likely contains date values."""
        # Sample the first few non-null values
        sample = series.dropna().head(5).astype(str)
        if len(sample) == 0:
            return False
            
        # Common date patterns
        date_patterns = [
            r'\d{4}-\d{1,2}-\d{1,2}',  # YYYY-MM-DD
            r'\d{1,2}/\d{1,2}/\d{2,4}',  # MM/DD/YYYY or DD/MM/YYYY
            r'\d{1,2}-\d{1,2}-\d{2,4}',  # MM-DD-YYYY or DD-MM-YYYY
            r'\d{4}/\d{1,2}/\d{1,2}'   # YYYY/MM/DD
        ]
        
        # Check if most values match date patterns
        matches = 0
        for val in sample:
            if any(re.search(pattern, str(val)) for pattern in date_patterns):
                matches += 1
                
        return matches / len(sample) > 0.5
    
    def _dataframe_to_markdown(self, df: pd.DataFrame, max_rows: int = 20) -> str:
        """Convert DataFrame to markdown table format."""
        if len(df) == 0 or len(df.columns) == 0:
            return "*Empty table*"
            
        # Limit rows for large tables
        if len(df) > max_rows:
            display_df = pd.concat([df.head(max_rows//2), df.tail(max_rows//2)])
            ellipsis_row = pd.DataFrame([{col: '...' for col in df.columns}])
            display_df = pd.concat([display_df.head(max_rows//2), ellipsis_row, display_df.tail(max_rows//2)])
        else:
            display_df = df
            
        # Create markdown table
        header = "| " + " | ".join(str(col) for col in display_df.columns) + " |"
        separator = "| " + " | ".join(["---" for _ in display_df.columns]) + " |"
        rows = ["| " + " | ".join(str(cell) for cell in row) + " |" for _, row in display_df.iterrows()]
        
        return "\n".join([header, separator] + rows)
    
    def _dataframe_to_html(self, df: pd.DataFrame, max_rows: int = 50) -> str:
        """Convert DataFrame to HTML table format."""
        try:
            # Limit rows for large tables
            if len(df) > max_rows:
                display_df = pd.concat([df.head(max_rows//2), df.tail(max_rows//2)])
            else:
                display_df = df
                
            # Use pandas HTML formatting
            html = display_df.to_html(index=False, na_rep="N/A", border=1)
            return html
        except Exception as e:
            return f"<p>Error generating HTML table: {str(e)}</p>"


class ChineseTheologyXMLLoader:
    """
    Custom loader for Chinese Theology XML files with specific tag structure:
    <title>, <author>, <periodical>, <content>
    
    This loader extracts structured data from XML and creates better document chunks
    with rich metadata for improved retrieval.
    """
    
    def __init__(self, file_path: str):
        """Initialize with path to XML file."""
        self.file_path = file_path
    
    def load(self) -> List[Document]:
        """
        Load and parse the XML file, returning a list of Documents.
        Each document contains content with appropriate metadata from XML tags.
        """
        try:
            # Parse the XML file
            tree = ET.parse(self.file_path)
            root = tree.getroot()
            
            documents = []
            
            # Extract common metadata that applies to all chunks
            title = self._get_tag_text(root, 'title', 'Untitled')
            author = self._get_tag_text(root, 'author', 'Unknown Author')
            periodical = self._get_tag_text(root, 'periodical', 'Unknown Publication')
            
            # Process the content tag which contains the main text
            content_elem = root.find('content')
            
            if content_elem is not None and content_elem.text:
                # Create the main document with full content
                main_doc = Document(
                    page_content=content_elem.text,
                    metadata={
                        'source': self.file_path,
                        'title': title,
                        'author': author,
                        'periodical': periodical,
                        'type': 'full_content'
                    }
                )
                documents.append(main_doc)
                
                # Also create a title document with rich metadata for better retrieval
                # This helps when users search for articles by title
                title_doc = Document(
                    page_content=f"Title: {title}\nAuthor: {author}\nPublication: {periodical}",
                    metadata={
                        'source': self.file_path,
                        'title': title,
                        'author': author,
                        'periodical': periodical,
                        'type': 'metadata'
                    }
                )
                documents.append(title_doc)
            
            # If there are section tags within content, create separate documents for each
            sections = root.findall('.//section')
            for i, section in enumerate(sections):
                if section.text and section.text.strip():
                    # Get section title if available
                    section_title = section.get('title', f'Section {i+1}')
                    
                    section_doc = Document(
                        page_content=section.text,
                        metadata={
                            'source': self.file_path,
                            'title': title,
                            'section': section_title,
                            'author': author,
                            'periodical': periodical,
                            'type': 'section'
                        }
                    )
                    documents.append(section_doc)
            
            # If no sections were found but we have paragraphs, use those
            if not sections:
                paragraphs = root.findall('.//p')
                for i, para in enumerate(paragraphs):
                    if para.text and para.text.strip():
                        para_doc = Document(
                            page_content=para.text,
                            metadata={
                                'source': self.file_path,
                                'title': title,
                                'paragraph': i+1,
                                'author': author,
                                'periodical': periodical,
                                'type': 'paragraph'
                            }
                        )
                        documents.append(para_doc)
            
            # If we didn't get any documents from sections or paragraphs,
            # and the main content is too long, split it into chunks
            if len(documents) <= 2 and content_elem is not None and content_elem.text:
                content_text = content_elem.text
                # Simple chunking by newlines if the content is long
                if len(content_text) > 2000:
                    chunks = content_text.split('\n\n')
                    for i, chunk in enumerate(chunks):
                        if chunk.strip():
                            chunk_doc = Document(
                                page_content=chunk,
                                metadata={
                                    'source': self.file_path,
                                    'title': title,
                                    'chunk': i+1,
                                    'total_chunks': len(chunks),
                                    'author': author,
                                    'periodical': periodical,
                                    'type': 'chunk'
                                }
                            )
                            documents.append(chunk_doc)
            
            print(f"Extracted {len(documents)} documents from {os.path.basename(self.file_path)}")
            return documents
            
        except Exception as e:
            print(f"Error parsing XML file {self.file_path}: {e}")
            # Return an empty document with error information
            return [Document(
                page_content=f"Error parsing XML file: {e}",
                metadata={'source': self.file_path, 'error': str(e)}
            )]
    
    def _get_tag_text(self, root, tag_name, default=""):
        """Helper to safely extract text from an XML tag."""
        element = root.find(tag_name)
        if element is not None and element.text:
            return element.text.strip()
        return default


class ChineseTheologyXMLDirectoryLoader:
    """
    Loads all XML files in a directory using the ChineseTheologyXMLLoader.
    """
    
    def __init__(self, directory_path: str, glob_pattern: str = "**/*.xml", recursive: bool = True):
        """Initialize with directory path and optional glob pattern."""
        self.directory_path = directory_path
        self.glob_pattern = glob_pattern
        self.recursive = recursive
    
    def load(self) -> List[Document]:
        """
        Load all XML files in the directory that match the glob pattern.
        """
        import glob
        
        documents = []
        
        # Generate file paths based on glob pattern
        if self.recursive:
            matches = glob.glob(os.path.join(self.directory_path, self.glob_pattern), recursive=True)
        else:
            matches = glob.glob(os.path.join(self.directory_path, self.glob_pattern))
        
        # Load each file
        for file_path in matches:
            if os.path.isfile(file_path) and file_path.lower().endswith('.xml'):
                try:
                    loader = ChineseTheologyXMLLoader(file_path)
                    file_docs = loader.load()
                    documents.extend(file_docs)
                except Exception as e:
                    print(f"Error loading {file_path}: {e}")
                    # Add an error document
                    documents.append(Document(
                        page_content=f"Error loading XML file: {e}",
                        metadata={'source': file_path, 'error': str(e)}
                    ))
        
        print(f"Loaded {len(documents)} total documents from {len(matches)} XML files")
        return documents
